import os
from openpyxl import Workbook
import configparser

current_path = os.getcwd()


def read_config():
    """
    从当前目录下面读取config.ini文件
    如果不存在就返回默认的字典{'rootpath': current_path, 'suffix': None, 'filename': 'default.xlsx'}
    存在则从配置文件中读取配置内容，并返回
    :return: dict
    """
    config_file = current_path + os.sep + "config.ini"
    if not os.path.exists(config_file):
        return {'rootpath': current_path, 'suffix': 'all', 'filename': 'default.xlsx'}
    else:
        config = configparser.ConfigParser()
        config.read(config_file, encoding='utf-8')
        return {'rootpath': config.get('fileinfo', 'root_dir'), 'suffix': config.get('fileinfo', 'dest_file_suffix_list').split(','), 'filename': config.get('fileinfo', 'save_file')}


def get_all_files(path, suffix):
    """
    获取指定路径下指定后缀名称的文件
    如果后缀名称为空，则读取全部
    :param path:
    :param suffix:
    :return:
    """
    file_list = list()
    for root, dirs, files in os.walk(path):
        for file in files:
            file_suffix = file.split('.')[-1]
            if suffix[0] != 'all':
                if file_suffix in suffix:
                    file_list.append(root + os.sep + file)
            else:
                file_list.append(root + os.sep + file)
    return file_list


def write_to_excel(content, filename):
    if not isinstance(content, list):
        raise ValueError("参数类型不正确")
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '文件名称'
    ws['B1'] = '文件路径'
    for i in range(len(content)):
        ws['A%d' % (i+2)] = content[i].split(os.sep)[-1]
        # 给单元格设置超链接
        ws['A%d' % (i + 2)].hyperlink = content[i]
        ws['A%d' % (i + 2)].style = "Hyperlink"
        ws['B%d' % (i + 2)] = content[i]
    wb.save(filename)


rc = read_config()
files = get_all_files(rc.get('rootpath'), suffix=rc.get('suffix'))
write_to_excel(files, rc.get('rootpath') + os.sep + rc.get('filename'))
